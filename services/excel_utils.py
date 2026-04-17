"""
Excel utilities for generating schedule templates and parsing.
Provides functionality to generate sample Excel files for non-technical users.
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def generate_schedule_template(output_path: str | None = None) -> bytes | None:
    """
    Generate a sample Excel template for schedule imports.
    
    Columns:
    - Группа (Group): Name of the student group
    - Опис (Weekday): Day of the week in Russian (Monday, Tuesday, etc.)
    - Дата (Date): Optional specific date for this override (YYYY-MM-DD)
    - Пара# (Lesson): Lesson number (1-5)
    - Предмет (Subject): Course subject
    - Преподаватель (Teacher): Teacher name
    - Кабинет (Room): Room number
    
    Args:
        output_path: If provided, saves to this file path. Otherwise returns bytes.
    
    Returns:
        bytes if output_path is None, otherwise None (writes to file)
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Расписание"
    
    # Define header columns
    headers = [
        "Группа",
        "День недели",
        "Дата (опционально)",
        "Пара №",
        "Предмет",
        "Преподаватель",
        "Кабинет",
        "Время начала (опционально)",
        "Время окончания (опционально)",
    ]
    
    # Add headers to first row
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set column widths
    column_widths = [12, 15, 15, 8, 20, 20, 10, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Add sample data rows
    sample_data = [
        ["10-ИТ-1", "Понедельник", "", 1, "Математика", "Иван Петров", "101", "09:00", "10:30"],
        ["10-ИТ-1", "Понедельник", "", 2, "Информатика", "Мария Сидорова", "202", "10:40", "12:10"],
        ["10-ИТ-1", "Вторник", "", 1, "Физика", "Петр Иванов", "103", "09:00", "10:30"],
        ["10-КБ-2", "Среда", "", 3, "Английский язык", "Анна Смирнова", "305", "12:20", "13:50"],
        ["10-КБ-2", "Четверг", "", 1, "История", "Виктор Кузнецов", "104", "09:00", "10:30"],
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Add instructions sheet
    instructions_sheet = workbook.create_sheet("Инструкция")
    instructions = [
        "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ РАСПИСАНИЯ",
        "",
        "Правила заполнения:",
        "1. Не удаляйте заголовки в первой строке",
        "2. Группа: Внесите точное название группы (например: 10-ИТ-1)",
        "3. День недели: Используйте дни на русском (Понедельник, Вторник, ...)",
        "4. Дата: Оставьте пустым для постоянного расписания, или введите дату в формате ГГГГ-ММ-ДД",
        "5. Пара №: Номер пары (1, 2, 3, 4, 5)",
        "6. Предмет: Название предмета/курса",
        "7. Преподаватель: ФИО преподавателя",
        "8. Кабинет: Номер кабинета",
        "9. Время: Заполняйте в формате ЧЧ:ММ (например: 09:00)",
        "",
        "Примечания:",
        "• Система игнорирует пустые строки и пробелы в начале/конце текста",
        "• Не чувствительна к регистру букв",
        "• Для замены расписания: пересохраните файл все новые данные",
        "",
        "Дни недели (допустимые значения):",
        "Понедельник, Вторник, Среда, Четверг, Пятница, Суббота",
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        cell = instructions_sheet.cell(row=row_num, column=1)
        cell.value = instruction
        if instruction.startswith("ИНСТРУКЦИЯ") or instruction.startswith("Правила"):
            cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    instructions_sheet.column_dimensions["A"].width = 80
    
    # Save or return
    if output_path:
        workbook.save(output_path)
        return None
    else:
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def generate_template_filename() -> str:
    """Generate a filename for the schedule template with current timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"schedule_template_{timestamp}.xlsx"
